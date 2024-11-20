import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from datetime import datetime

def statistic(request):
    # Khởi tạo Workbook
    wb = openpyxl.Workbook()
    
    # Tạo các sheet
    daily_sheet = wb.active
    daily_sheet.title = "Thống kê theo ngày"
    monthly_sheet = wb.create_sheet(title="Thống kê theo tháng")
    yearly_sheet = wb.create_sheet(title="Thống kê theo năm")

    # Hàm ghi dữ liệu vào sheet
    def write_to_sheet(sheet, data, headers):
        # Thêm tiêu đề
        sheet.append(headers)
        for header in sheet[1]:
            header.font = Font(bold=True)
        
        # Thêm dữ liệu
        for item in data:
            sheet.append(list(item.values()))

    # Lấy dữ liệu thống kê
    statistics = request.GET.get('statistics', [])
    monthly_statistics = request.GET.get('monthly_statistics', [])
    yearly_statistics = request.GET.get('yearly_statistics', [])
    
    # Ghi dữ liệu vào các sheet
    daily_headers = ["STT", "Ngày", "Chi phí", "Doanh thu", "Lợi nhuận", "%Tăng trưởng theo ngày"]
    write_to_sheet(daily_sheet, statistics, daily_headers)
    
    monthly_headers = ["STT", "Tháng", "Chi phí", "Doanh thu", "Lợi nhuận", "%Tăng trưởng theo tháng"]
    write_to_sheet(monthly_sheet, monthly_statistics, monthly_headers)
    
    yearly_headers = ["STT", "Năm", "Chi phí", "Doanh thu", "Lợi nhuận", "%Tăng trưởng theo năm"]
    write_to_sheet(yearly_sheet, yearly_statistics, yearly_headers)

    # Tên file xuất
    filename = f"Thong_Ke_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Lưu workbook ra HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename={filename}"
    wb.save(response)
    return response
