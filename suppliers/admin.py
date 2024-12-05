from django.contrib import admin
import openpyxl
from django.http import HttpResponse

from suppliers.forms import StockEntryForm, SupplierForm
from .models import Supplier, StockEntry
from django.db.models import Sum
from django.utils.timezone import now
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from django.db.models import F, ExpressionWrapper, DecimalField
from openpyxl.styles import Alignment
# Register your models here.

def export_daily_monthly_yearly_costs_to_excel(modeladmin, request, queryset):
    current_year = now().year  # Lấy năm hiện tại để đặt tên file

    # Tính tổng chi phí theo ngày, tháng, năm
    def calculate_total_data(trunc_func):
        return (
            queryset
            .annotate(period=trunc_func('entry_date'))  # Truncation theo ngày, tháng, năm
            .annotate(
                cost_per_entry=ExpressionWrapper(
                    (F('quantity') - F('remaining_quantity')) * F('unit_price'),
                    output_field=DecimalField()
                )
            )
            .values('period')
            .annotate(total_cost=Sum('cost_per_entry'))  # Tổng chi phí cho từng khoảng thời gian
        )

    daily_data = calculate_total_data(TruncDay)
    monthly_data = calculate_total_data(TruncMonth)
    yearly_data = calculate_total_data(TruncYear)

    # Tạo workbook mới
    workbook = openpyxl.Workbook()
    
    def populate_sheet(sheet, data, period_format, title):
        sheet.title = title
        sheet.append([period_format, 'Tổng chi tiêu'])  # Thêm tiêu đề cột

        total_cost = 0
        
        for entry in data:
            # Định dạng ngày, tháng, năm tương ứng
            formatted_period = entry['period'].strftime('%d-%m-%Y') if title == 'Chi tiêu theo ngày' else (
                entry['period'].strftime('%m-%Y') if title == 'Chi tiêu theo tháng' else entry['period'].strftime('%Y')
            )
            total_cost += entry['total_cost']

        sheet.append([formatted_period, total_cost])
        
        sheet.cell(row=2, column=2).number_format = '#,##0'
        sheet.cell(row=2, column=2).alignment = Alignment(horizontal='right')

    # Tạo và định dạng sheet cho từng trường hợp
    day_sheet = workbook.active
    populate_sheet(day_sheet, daily_data, 'Ngày', 'Chi tiêu theo ngày')
    month_sheet = workbook.create_sheet()
    populate_sheet(month_sheet, monthly_data, 'Tháng', 'Chi tiêu theo tháng')
    year_sheet = workbook.create_sheet()
    populate_sheet(year_sheet, yearly_data, 'Năm', 'Chi tiêu theo năm')

    # Trả về file Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=costs_{current_year}.xlsx'
    workbook.save(response)
    return response

export_daily_monthly_yearly_costs_to_excel.short_description = "Xuất Chi Phí Ra Excel"
class SupplierAdmin(admin.ModelAdmin):
    form = SupplierForm
    list_display = ['supplier_name', 'supplier_email', 'supplier_phone', 'supplier_country','supplier_city', 'supplier_is_active']
    search_fields = ['supplier_name', 'supplier_email']
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        form.save_m2m()
class StockEntryAdmin(admin.ModelAdmin):
    form = StockEntryForm
    fields = ['product', 'supplier', 'quantity','total_value', 'unit_price','stock_category','stock_color','stock_value','stock_size']
    list_display = ['product', 'stock_color','stock_size', 'supplier', 'quantity','remaining_quantity','formatted_total_value', 'formatted_unit_price','entry_date']  # Hiển thị thông tin trong danh sách
    readonly_fields = ['total_value']
    search_fields = ['supplier__supplier_name']
    actions = [export_daily_monthly_yearly_costs_to_excel]  # Thêm action vào admin
    def has_delete_permission(self, request, obj=None):
        return True
    def formatted_total_value(self, obj):
        return "{:,.0f}".format(obj.total_value)  # Định dạng với dấu phẩy
    formatted_total_value.short_description = 'Tổng giá'
    def formatted_unit_price(self, obj):
        return "{:,.0f}".format(obj.unit_price)  # Định dạng với dấu phẩy
    formatted_unit_price.short_description = 'Đơn giá'

admin.site.register(StockEntry, StockEntryAdmin)
admin.site.register(Supplier, SupplierAdmin)