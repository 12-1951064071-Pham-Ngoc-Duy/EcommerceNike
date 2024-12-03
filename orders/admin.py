from django.contrib import admin
from orders.forms import OrderFormAdmin
from suppliers.models import StockEntry
from .models import Payment, Order, OrderProduct
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Sum,F
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from datetime import timezone
from decimal import Decimal

def export_profit_to_excel(modeladmin, request, queryset):
    workbook = Workbook()

    # --- Sheet 1: Lợi Nhuận Hàng Ngày ---
    sheet1 = workbook.active
    sheet1.title = "Lợi Nhuận Ngày"
    sheet1.append(['Ngày', 'Doanh Thu', 'Chi Phí', 'Lợi Nhuận', '% Tăng Trưởng'])

    doanh_thu_ngay = (
        Order.objects.filter(order_is_ordered=True)
        .annotate(ngay=TruncDay('order_created_at'))
        .values('ngay')
        .annotate(doanh_thu=Sum('order_total'))
    )

    chi_phi_ngay = (
        StockEntry.objects.annotate(ngay=TruncDay('entry_date'))
        .values('ngay')
        .annotate(chi_phi=Sum((F('quantity') - F('remaining_quantity')) * F('unit_price')))
    )

    previous_day_revenue = None
    for doanh_thu in doanh_thu_ngay:
        ngay = doanh_thu['ngay'].astimezone(timezone.utc).replace(tzinfo=None)
        doanh_thu_value = Decimal(doanh_thu['doanh_thu'])

        # Lấy chi phí tương ứng với ngày
        chi_phi = next(
            (cp['chi_phi'] for cp in chi_phi_ngay if cp['ngay'].date() == ngay.date()),
            Decimal(0),
        )
        loi_nhuan = doanh_thu_value - chi_phi

        # Tính % tăng trưởng theo ngày
        growth_rate = Decimal(0)
        if previous_day_revenue is not None and previous_day_revenue > 0:
            growth_rate = ((doanh_thu_value - previous_day_revenue) / previous_day_revenue) * 100

        sheet1.append([ngay, doanh_thu_value, chi_phi, loi_nhuan, f"{growth_rate:.2f}%"])
        previous_day_revenue = doanh_thu_value

    # --- Sheet 2: Lợi Nhuận Hàng Tháng ---
    sheet2 = workbook.create_sheet(title="Lợi Nhuận Tháng")
    sheet2.append(['Tháng', 'Doanh Thu', 'Chi Phí', 'Lợi Nhuận', '% Tăng Trưởng'])

    doanh_thu_thang = (
        Order.objects.filter(order_is_ordered=True)
        .annotate(thang=TruncMonth('order_created_at'))
        .values('thang')
        .annotate(doanh_thu=Sum('order_total'))
    )

    chi_phi_thang = (
        StockEntry.objects.annotate(thang=TruncMonth('entry_date'))
        .values('thang')
        .annotate(chi_phi=Sum((F('quantity') - F('remaining_quantity')) * F('unit_price')))
    )

    previous_month_revenue = None
    for doanh_thu in doanh_thu_thang:
        thang = doanh_thu['thang'].month
        doanh_thu_value = Decimal(doanh_thu['doanh_thu'])

        chi_phi = next(
            (cp['chi_phi'] for cp in chi_phi_thang if cp['thang'].month == thang),
            Decimal(0),
        )
        loi_nhuan = doanh_thu_value - chi_phi

        # Tính % tăng trưởng theo tháng
        growth_rate = Decimal(0)
        if previous_month_revenue is not None and previous_month_revenue > 0:
            growth_rate = ((doanh_thu_value - previous_month_revenue) / previous_month_revenue) * 100

        sheet2.append([thang, doanh_thu_value, chi_phi, loi_nhuan, f"{growth_rate:.2f}%"])
        previous_month_revenue = doanh_thu_value

    # --- Sheet 3: Lợi Nhuận Hàng Năm ---
    sheet3 = workbook.create_sheet(title="Lợi Nhuận Năm")
    sheet3.append(['Năm', 'Doanh Thu', 'Chi Phí', 'Lợi Nhuận', '% Tăng Trưởng'])

    doanh_thu_nam = (
        Order.objects.filter(order_is_ordered=True)
        .annotate(nam=TruncYear('order_created_at'))
        .values('nam')
        .annotate(doanh_thu=Sum('order_total'))
    )

    chi_phi_nam = (
        StockEntry.objects.annotate(nam=TruncYear('entry_date'))
        .values('nam')
        .annotate(chi_phi=Sum((F('quantity') - F('remaining_quantity')) * F('unit_price')))
    )


    previous_year_revenue = None
    for doanh_thu in doanh_thu_nam:
        nam = doanh_thu['nam'].year
        doanh_thu_value = Decimal(doanh_thu['doanh_thu'])

        chi_phi = next(
            (cp['chi_phi'] for cp in chi_phi_nam if cp['nam'].year == nam),
            Decimal(0),
        )
        loi_nhuan = doanh_thu_value - chi_phi

        # Tính % tăng trưởng theo năm
        growth_rate = Decimal(0)
        if previous_year_revenue is not None and previous_year_revenue > 0:
            growth_rate = ((doanh_thu_value - previous_year_revenue) / previous_year_revenue) * 100

        sheet3.append([nam, doanh_thu_value, chi_phi, loi_nhuan, f"{growth_rate:.2f}%"])
        previous_year_revenue = doanh_thu_value

    # Tạo response trả về file Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="loi_nhuan.xlsx"'

    workbook.save(response)
    return response

export_profit_to_excel.short_description = "Xuất Lợi Nhuận Ra Excel"


class OrderProductInline(admin.TabularInline):
    model = OrderProduct
    readonly_fields = ('payment', 'user', 'product', 'order_product_quantity', 'formatted_order_product_price', 'order_product_ordered')
    list_display = ('payment', 'user', 'product', 'order_product_quantity', 'formatted_order_product_price', 'order_product_ordered')
    extra = 0
    def formatted_order_product_price(self, obj):
        return "{:,.0f}".format(obj.order_product_price)  # Định dạng với dấu phẩy
    formatted_order_product_price.short_description = 'Gía'
    def has_add_permission(self, request, obj=None):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return request.user.is_staff


class OrderAdmin(admin.ModelAdmin):
    form = OrderFormAdmin
    list_display = ['order_number', 'full_name', 'order_phone', 'order_email', 'order_country', 'order_city', 'formatted_order_total', 'order_village', 'formatted_order_tax', 'order_status', 'order_created_at', 'order_is_ordered']
    list_filter = ['order_status', 'order_is_ordered']
    search_fields = ['order_number', 'payment__payment_id']
    list_per_page = 20
    inlines = [OrderProductInline]
    actions = [export_profit_to_excel]

    def formatted_order_total(self, obj):
        return "{:,.0f}".format(obj.order_total)  # Định dạng với dấu phẩy
    formatted_order_total.short_description = 'Tổng'

    def formatted_order_tax(self, obj):
        return "{:,.0f}".format(obj.order_tax)  # Định dạng với dấu phẩy
    formatted_order_tax.short_description = 'Phí giao hàng'

    def get_total_cost(self, obj):
        total_cost = StockEntry.objects.filter(entry_date__year=obj.order_created_at.year).aggregate(
            total=Sum('total_value')
        )['total'] or Decimal(0)  # Đảm bảo giá trị trả về là Decimal
        return total_cost

    def get_profit(self, obj):
        total_cost = self.get_total_cost(obj)
        return Decimal(obj.order_total) - total_cost
    
    get_total_cost.short_description = 'Chi Phí'
    get_profit.short_description = 'Lợi Nhuận'
    def get_readonly_fields(self, request, obj=None):
        # Nếu đang chỉnh sửa (không phải tạo mới), chỉ cho phép thay đổi order_status
        if obj:
            return [field.name for field in self.model._meta.fields if field.name != 'order_status' and field.name != 'order_is_ordered']
        return super().get_readonly_fields(request, obj)

    def has_add_permission(self, request, obj=None):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return False

    def has_view_permission(self, request, obj=None):
        return request.user.is_staff

admin.site.register(Order, OrderAdmin)